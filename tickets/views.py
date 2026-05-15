import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
from django.http import HttpResponse
from rest_framework import generics, filters
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from django_filters.rest_framework import DjangoFilterBackend

from .models import Ticket, Category, TicketActivity, Notification
from .serializers import (
    TicketListSerializer, TicketDetailSerializer,
    TicketCreateSerializer, CategorySerializer,
    TicketCommentSerializer, TimeLogSerializer,
    TicketActivitySerializer, NotificationSerializer,
)
from .filters import TicketFilter
from .pagination import TicketPagination
from . import activity as act
from . import notifications as notif


class CategoryListView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = CategorySerializer
    queryset = Category.objects.filter(is_active=True)


class TicketListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    pagination_class = TicketPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = TicketFilter
    search_fields = ['title', 'description']
    ordering_fields = ['created_at', 'updated_at', 'priority']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return TicketCreateSerializer
        return TicketListSerializer

    def get_queryset(self):
        user = self.request.user
        qs = Ticket.objects.select_related('company', 'assigned_to', 'created_by', 'category')
        if user.role in ('admin', 'dispatcher'):
            return qs
        if user.role == 'technician':
            return qs.filter(assigned_to=user)
        if user.role == 'client':
            return qs.filter(company=user.company)
        return qs.none()


class TicketDetailView(generics.RetrieveUpdateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = TicketDetailSerializer

    def get_queryset(self):
        user = self.request.user
        qs = Ticket.objects.select_related('company', 'assigned_to', 'created_by', 'category')
        if user.role in ('admin', 'dispatcher'):
            return qs
        if user.role == 'technician':
            return qs.filter(assigned_to=user)
        if user.role == 'client':
            return qs.filter(company=user.company)
        return qs.none()

    def perform_update(self, serializer):
        ticket = self.get_object()
        old_data = {
            'status': ticket.status,
            'assigned_to_id': ticket.assigned_to_id,
            'priority': ticket.priority,
        }
        instance = serializer.save()
        instance.refresh_from_db(fields=['assigned_to'])
        act.log_changes(instance, self.request.user, old_data)
        if old_data['assigned_to_id'] != instance.assigned_to_id:
            notif.on_ticket_assigned(instance, self.request.user)
        if old_data['status'] != instance.status and instance.status == 'resolved':
            notif.on_ticket_resolved(instance, self.request.user)


class TicketActivityListView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = TicketActivitySerializer

    def get_queryset(self):
        return TicketActivity.objects.filter(
            ticket_id=self.kwargs['pk']
        ).select_related('actor').order_by('created_at')


class TicketCommentListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = TicketCommentSerializer

    def get_queryset(self):
        qs = Ticket.objects.get(pk=self.kwargs['pk']).comments.select_related('author')
        user = self.request.user
        if user.role == 'client':
            qs = qs.filter(is_internal=False)
        return qs

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['ticket_id'] = self.kwargs['pk']
        return ctx

    def perform_create(self, serializer):
        instance = serializer.save()
        act.log_comment(instance.ticket, self.request.user)
        notif.on_comment_added(instance.ticket, self.request.user, instance.body)


class TimeLogListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = TimeLogSerializer

    def get_queryset(self):
        return Ticket.objects.get(pk=self.kwargs['pk']).timelogs.select_related('technician')

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['ticket_id'] = self.kwargs['pk']
        return ctx


class TicketExportView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = TicketFilter
    search_fields = ['title', 'description']
    ordering = ['-created_at']

    def get_queryset(self):
        user = self.request.user
        qs = Ticket.objects.select_related('company', 'assigned_to', 'created_by', 'category')
        if user.role in ('admin', 'dispatcher'):
            return qs
        if user.role == 'technician':
            return qs.filter(assigned_to=user)
        if user.role == 'client':
            return qs.filter(company=user.company)
        return qs.none()

    def get(self, request):
        qs = self.filter_queryset(self.get_queryset())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Tickets'

        headers = ['#', 'Title', 'Status', 'Priority', 'Category', 'Company',
                   'Assigned To', 'Created By', 'Created At', 'Updated At',
                   'SLA Breach', 'Due Date']
        ws.append(headers)

        header_fill = PatternFill('solid', fgColor='4F46E5')
        header_font = Font(bold=True, color='FFFFFF')
        thin = Side(style='thin', color='D1D5DB')
        border = Border(bottom=thin)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border

        def name(user_obj):
            if not user_obj:
                return ''
            n = f'{user_obj.first_name} {user_obj.last_name}'.strip()
            return n or user_obj.username

        def dt(value):
            if not value:
                return ''
            return value.strftime('%Y-%m-%d %H:%M')

        STATUS_MAP = {'new': 'New', 'assigned': 'Assigned', 'in_progress': 'In Progress',
                      'on_hold': 'On Hold', 'resolved': 'Resolved', 'closed': 'Closed'}
        PRIORITY_MAP = {'low': 'Low', 'medium': 'Medium', 'high': 'High', 'critical': 'Critical'}

        for ticket in qs:
            ws.append([
                ticket.id,
                ticket.title,
                STATUS_MAP.get(ticket.status, ticket.status),
                PRIORITY_MAP.get(ticket.priority, ticket.priority),
                ticket.category.name if ticket.category else '',
                ticket.company.name,
                name(ticket.assigned_to),
                name(ticket.created_by),
                dt(ticket.created_at),
                dt(ticket.updated_at),
                'Yes' if ticket.sla_breach else 'No',
                dt(ticket.due_date),
            ])

        col_widths = [6, 45, 12, 10, 16, 20, 22, 22, 17, 17, 10, 17]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        ws.row_dimensions[1].height = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="tickets.xlsx"'
        wb.save(response)
        return response


class NotificationListView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = NotificationSerializer
    pagination_class = None

    def get_queryset(self):
        return (
            Notification.objects
            .filter(user=self.request.user)
            .select_related('ticket')
            .order_by('-created_at')[:20]
        )


class NotificationMarkAllReadView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        return Response({'status': 'ok'})
